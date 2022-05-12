import csv
import email
import mimetypes

from pip._vendor import requests
from openpyxl.reader.excel import load_workbook
from datetime import datetime, timedelta


# Diccionarios que contienen los dias y meses en español 
daysDic = {
            "Mon":'Lunes', "Tue":'Martes', "Wed":'Miércoles', "Thu":'Jueves', "Fri":'Viernes', "Sat":'Sábado', "Sun":'Domingo'
        }
monthsDic = {
            "1":'Enero', "2":'Febrero', "3":'Marzo', "4":'Abril', "5":'Mayo', "6":'Junio', "7":'Julio', "8":'Agosto', "9":'Septiembre', "10":'Octubre', "11":'Noviembre', "12":'Diciembre'
        }
#

# Funcion de filtración de datos
def filter_row(row, headers):

    # Apartir de la fecha actual (today()), obtenemos el resto de las fechas que se necesitan para condicionar y poder filtrar los resultados
    today = datetime.today()
    today_dt = today.date()
    prev_15days_dt = (today - timedelta(days=15)).date()
    this_week_dt = (today + timedelta(days=5)).date()
    next_10days_dt = (today + timedelta(days=10)).date()
    next_15days_dt = (today + timedelta(days=15)).date()

    # Creamos un diccionario con los headers obtenidos durante la lectura del CSV, y cada fila (row) que se recorre de dicho archivo
    dic = dict(zip(headers, row))


    if headers[1] == 'fechaInic' and headers[2] == 'fechaTerm':

        name = dic['nombre_completo']
        ini_dt = datetime.strptime(dic['fechaInic'], '%Y-%m-%d %H:%M:%S').date()
        ter_dt = datetime.strptime(dic['fechaTerm'], '%Y-%m-%d %H:%M:%S').date()

        ini_day_word = ini_dt.strftime("%a")
        ini_day = daysDic[ini_day_word][:3]
        ter_day_word = ter_dt.strftime("%a")
        ter_day = daysDic[ter_day_word][:3]

        ini_month_word = ini_dt.month
        ini_month = monthsDic[str(ini_month_word)][:3]
        ter_month_word = ter_dt.month
        ter_month = monthsDic[str(ter_month_word)][:3]

        if ini_dt > prev_15days_dt and ini_dt < today_dt :

            if ter_dt == today_dt or (ter_dt > today_dt and ter_dt < next_15days_dt) or (ter_dt > today_dt and ter_dt > next_15days_dt): 
                return ('ausentes', f'{name.title()}', 'fechas', f'{ter_day} {ter_dt.day} de {ter_month}')


        if ini_dt >= today_dt and ini_dt < next_15days_dt :

            if ini_dt == today_dt : 
                return ('ausentes', f'{name.title()}', 'fechas', f'{ter_day} {ter_dt.day} de {ter_month}')

            if (ini_dt > today_dt and ter_dt < next_15days_dt) or (ini_dt > today_dt and ter_dt > next_15days_dt) :
                return ('proximos', f'{name.title()}', 'fechas', f'{ini_day} {ini_dt.day} de {ini_month} al {ter_day} {ter_dt.day} de {ter_month}')


    if headers[1] == 'fechaInicio' and headers[2] == 'fechaTermino':

        name = dic['nombre_completo']
        ini_dt = datetime.strptime(dic['fechaInicio'], '%Y-%m-%d %H:%M:%S').date()
        ter_dt = datetime.strptime(dic['fechaTermino'], '%Y-%m-%d %H:%M:%S').date()

        ini_day_word = ini_dt.strftime("%a")
        ini_day = daysDic[ini_day_word][:3]
        ter_day_word = ter_dt.strftime("%a")
        ter_day = daysDic[ter_day_word][:3]

        ini_month_word = ini_dt.month
        ini_month = monthsDic[str(ini_month_word)][:3]
        ter_month_word = ter_dt.month
        ter_month = monthsDic[str(ter_month_word)][:3]

        if ini_dt > prev_15days_dt and ini_dt < today_dt:

            if ter_dt == today_dt or ter_dt > today_dt :
                return ('ausentes', f'{name.title()}', 'fechas', f'{ter_day} {ter_dt.day} de {ter_month}')

        if ini_dt == today_dt :

            if ter_dt == ini_dt :
                return ('ausentes', f'{name.title()}', 'fechas', f'Sólo hoy {ini_day} {ini_dt.day} de {ini_month}')

            if ter_dt > today_dt and ter_dt <= this_week_dt :
                return ('ausentes', f'{name.title()}', 'fechas', 'fechas', f'{ini_day} {ini_dt.day} de {ini_month} al {ter_day} {ter_dt.day} de {ter_month}')

        if ini_dt > today_dt and ini_dt <= this_week_dt :

            if ter_dt == ini_dt :
                return ('proximos', f'{name.title()}', 'fechas', f'Sólo el {ini_day} {ini_dt.day} de {ini_month}')           
            else:
                if (ter_dt > ini_dt  and ter_dt <= this_week_dt) or (ter_dt > this_week_dt and ter_dt < next_10days_dt) :
                    return ('proximos', f'{name.title()}', 'fechas', f'{ini_day} {ini_dt.day} de {ini_month} al {ter_day} {ter_dt.day} de {ter_month}')
            

    return (None,None,None,None)


# Funcion para obtener las listas con sus respectivos datos
# Listas de ausentes
def absent_lists(csv_reader_vac, headers):
    
    absent_names, absent_dates = [], []
    next_absent_names, next_absent_dates = [], []

    for row in csv_reader_vac:

        if '' not in row:
            category, fr_names, dates, fr_dates = filter_row(row, headers)
            if fr_names and fr_dates:
                if category == 'ausentes' and dates == 'fechas':
                    absent_names.append(fr_names) 
                    absent_dates.append(fr_dates)                    
                if category == 'proximos' and dates == 'fechas':
                    next_absent_names.append(fr_names) 
                    next_absent_dates.append(fr_dates)
    
    return absent_names, absent_dates, next_absent_names, next_absent_dates

def main(attached_file):

    msg = email.message_from_bytes(attached_file)

    for i, part in enumerate(msg.walk(), 1):

        if i == 3:
            filename0='/tmp/vacaciones.xlsx'

            with open(filename0, 'wb') as fv:
                fv.write(part.get_payload(decode = True))

        if i == 4:
            filename1='/tmp/permisos.xlsx'

            with open(filename1, 'wb') as fp:
                fp.write(part.get_payload(decode=True))


    #Extraccion de datos VACACIONES
    wb = load_workbook(filename='/tmp/vacaciones.xlsx')
    sheet = wb.active
    csv_data = []

    for value in sheet.iter_rows(values_only=True):
        csv_data.append(list(value))

    with open('/tmp/consultaVacaciones.csv', 'w') as csv_file_vac:
        csv_writer = csv.writer(csv_file_vac, delimiter = ";")

        for line in csv_data:
            csv_writer.writerow(line)
    #Fin extracción datos V.

    #Extraccion de datos PERMISOS ADMINISTRATIVOS
    wb = load_workbook(filename='/tmp/permisos.xlsx')
    sheet = wb.active
    csv_data = []

    for value in sheet.iter_rows(values_only=True):
        csv_data.append(list(value))

    with open('/tmp/consultaPermisos.csv', 'w') as csv_file_perm:
        csv_writer = csv.writer(csv_file_perm, delimiter = ";")

        for line in csv_data:
            csv_writer.writerow(line)
    #Fin extracción datos P.A.

    with open('/tmp/consultaVacaciones.csv','r') as csv_file_vac:
        csv_reader_vac = csv.reader(csv_file_vac, delimiter = ';')
        
        next(csv_reader_vac)
        headers = next(csv_reader_vac)

        absent_name_1, absent_date_1, next_absent_name_1, next_absent_date_1 = absent_lists(csv_reader_vac, headers)


    with open('/tmp/consultaPermisos.csv','r') as csv_file_perm:
        csv_reader_perm = csv.reader(csv_file_perm, delimiter = ';')
        
        next(csv_reader_perm)
        headers = next(csv_reader_perm)

        absent_name_2, absent_date_2, next_absent_name_2, next_absent_date_2 = absent_lists(csv_reader_perm, headers)


    absent_name_1.extend(absent_name_2)
    absent_date_1.extend(absent_date_2)
    next_absent_name_1.extend(next_absent_name_2)
    next_absent_date_1.extend(next_absent_date_2)

    if not absent_name_1 and not absent_date_1 :    
        absent_name_1.append("...")
        absent_date_1.append("...")

    if not next_absent_name_1 and not next_absent_date_1 :
        next_absent_name_1.append("...")
        next_absent_date_1.append("...")

    
    # Webhook del canal receptor del mensaje
    # general
    #url = "https://hooks.slack.com/services/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"
    # rodrigo
    url = "https://hooks.slack.com/services/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"

    absent_name_str = '\n'.join(absent_name_1)
    absent_date_str = '\n'.join(absent_date_1)
    next_absent_name_str = '\n'.join(next_absent_name_1)
    next_absent_date_str = '\n'.join(next_absent_date_1)


    result = requests.post(url, json = {'text': "Vacaciones Unholster",
        "blocks":
            [
                {
                    "type": "header",
                    "text": {
                        "type": "plain_text",
                        "text": "Unholsterianos Ausentes!\n\n"
                    }
                },
                {
                    "type": "divider"
                },
                {
                    "type": "section",
                    "fields": [
                        {
                            "type": "mrkdwn",
                            "text": f"*Ausentes *\n{absent_name_str}"
                        },
                        {
                            "type": "mrkdwn",
                            "text": f"*Término*\n{absent_date_str}"
                        }
                    ]       
                },
                {
                    "type": "divider"
                },
                {
                    "type": "section",
                    "fields": [
                        {
                            "type": "mrkdwn",
                            "text": f"*Próximamente*\n{next_absent_name_str}"
                        },
                        {
                            "type": "mrkdwn",
                            "text": f"*Inicio - Término*\n{next_absent_date_str}"
                        }
                    ]       
                },
                {
                    "type": "divider"
                }
            ]
        }
    )
